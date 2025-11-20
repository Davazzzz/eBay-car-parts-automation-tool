from flask import Flask, render_template, request, jsonify
from config import Config
from junkyard_prices import JunkyardPrices
from ebay_api import EbayAPI
from parts_analyzer import PartsAnalyzer
from saved_parts import SavedPartsList
from ebay_link_parser import EbayLinkParser

app = Flask(__name__)
app.config.from_object(Config)

# Initialize components
junkyard_prices = JunkyardPrices(Config.JUNKYARD_PRICES_CSV)
ebay_api = EbayAPI()
analyzer = PartsAnalyzer(ebay_api, junkyard_prices)
saved_list = SavedPartsList(Config.SAVED_PARTS_DB)
link_parser = EbayLinkParser()

@app.route('/')
def index():
    """Home page"""
    return render_template('index.html')

@app.route('/analyze', methods=['POST'])
def analyze():
    """Analyze vehicle parts"""
    try:
        data = request.json

        year = data.get('year')
        make = data.get('make')
        model = data.get('model')
        trim = data.get('trim', '')
        vehicle_type = data.get('vehicle_type', 'car')
        filter_type = data.get('filter_type', 'high_priority')  # high_priority, light, or all

        # Analyze vehicle
        results = analyzer.analyze_vehicle(year, make, model, vehicle_type, filter_type)

        # Get top 5 parts
        top_parts = analyzer.get_top_parts(results, 5)

        # Generate summary
        summary = {
            'total_parts': len(results),
            'high_roi_count': len([r for r in results if r['roi_rating'] == 'High']),
            'top_5_parts': [
                {'name': p['part_name'], 'roi': p['roi']}
                for p in top_parts
            ],
            'vehicle_info': f"{year} {make} {model} {trim}".strip()
        }

        return jsonify({
            'success': True,
            'results': results,
            'summary': summary
        })
    except Exception as e:
        print(f"ERROR in /analyze: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/save_part', methods=['POST'])
def save_part():
    """Save a part to the list"""
    data = request.json
    saved_list.add_part(data)

    return jsonify({
        'success': True,
        'message': f"Saved: {data['part_name']}"
    })

@app.route('/manual_add', methods=['POST'])
def manual_add():
    """Manually add a part"""
    data = request.json

    part_name = data.get('part_name')
    junkyard_price = float(data.get('junkyard_price'))
    ebay_sold_price = float(data.get('ebay_sold_price'))

    part_data = saved_list.add_manual(part_name, junkyard_price, ebay_sold_price)

    return jsonify({
        'success': True,
        'part': part_data
    })

@app.route('/add_from_link', methods=['POST'])
def add_from_link():
    """Add a part from eBay listing link"""
    try:
        data = request.json
        ebay_url = data.get('ebay_url')
        custom_part_name = data.get('custom_part_name', '').strip()
        vehicle_type = data.get('vehicle_type', 'car')  # car or truck
        selected_junkyard_parts = data.get('junkyard_parts', [])  # List of selected junkyard parts
        youtube_link = data.get('youtube_link', '')
        notes = data.get('notes', '')

        # Parse the eBay link
        parsed_data = link_parser.parse_link(ebay_url)

        if not parsed_data['success']:
            return jsonify({
                'success': False,
                'error': parsed_data.get('error', 'Failed to parse eBay link')
            }), 400

        # Determine part name: custom > junkyard part > auto-extracted from eBay
        if custom_part_name:
            # Use custom part name if provided
            part_name = custom_part_name
        elif selected_junkyard_parts:
            # Use first junkyard part name if junkyard parts were selected
            part_name = selected_junkyard_parts[0]
        else:
            # Extract part name from eBay title as fallback
            part_name = link_parser.extract_part_name(parsed_data['title'])

        # Calculate junkyard price from selected parts
        junkyard_price = 0.0
        junkyard_part_names = []

        if selected_junkyard_parts:
            # User manually selected junkyard parts
            for junk_part_name in selected_junkyard_parts:
                price = junkyard_prices.get_price(junk_part_name)
                if price:
                    junkyard_price += price
                    junkyard_part_names.append(junk_part_name)

        # If no parts selected, try auto-matching
        if junkyard_price == 0:
            # Try to find junkyard price
            matched_price = junkyard_prices.get_price(part_name)

            # If not found, try to match keywords
            if matched_price is None:
                all_parts = junkyard_prices.get_all_parts()
                part_name_upper = part_name.upper()
                for junk_part in all_parts:
                    if part_name_upper in junk_part.upper() or junk_part.upper() in part_name_upper:
                        matched_price = junkyard_prices.get_price(junk_part)
                        if matched_price:
                            junkyard_part_names = [junk_part]
                            break

            if matched_price:
                junkyard_price = matched_price

        # Calculate ROI
        ebay_price = parsed_data['price']
        roi = 0
        if junkyard_price > 0 and ebay_price > 0:
            roi = ebay_price / junkyard_price

        # Determine ROI rating
        if roi < 2:
            roi_rating = "Low"
        elif roi < 5:
            roi_rating = "Medium"
        else:
            roi_rating = "High"

        # Create part data
        part_data = {
            'part_name': part_name,
            'ebay_title': parsed_data['title'],
            'ebay_url': ebay_url,
            'ebay_price': ebay_price,
            'junkyard_price': junkyard_price,
            'junkyard_parts': junkyard_part_names,  # List of junkyard parts used
            'roi': roi,
            'roi_rating': roi_rating,
            'vehicle_type': vehicle_type,
            'year': parsed_data.get('year', ''),
            'make': parsed_data.get('make', ''),
            'model': parsed_data.get('model', ''),
            'youtube_link': youtube_link,
            'notes': notes
        }

        # Add to saved list
        saved_list.add_part(part_data)

        return jsonify({
            'success': True,
            'part': part_data
        })

    except Exception as e:
        print(f"ERROR in /add_from_link: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/junkyard_parts')
def get_junkyard_parts():
    """Get all junkyard parts with prices"""
    all_parts = junkyard_prices.get_all_parts()
    parts_with_prices = []

    for part in all_parts:
        price = junkyard_prices.get_price(part)
        if price:
            parts_with_prices.append({
                'name': part,
                'price': price
            })

    return jsonify({
        'success': True,
        'parts': parts_with_prices
    })

@app.route('/saved_list')
def get_saved_list():
    """Get all saved parts"""
    parts = saved_list.get_all()

    return jsonify({
        'success': True,
        'parts': parts
    })

@app.route('/remove_part/<int:index>', methods=['DELETE'])
def remove_part(index):
    """Remove a part from saved list"""
    success = saved_list.remove_part(index)

    return jsonify({
        'success': success
    })

@app.route('/update_part/<int:index>', methods=['PUT'])
def update_part(index):
    """Update a saved part's notes and YouTube link"""
    try:
        data = request.json
        youtube_link = data.get('youtube_link', '')
        notes = data.get('notes', '')

        if 0 <= index < len(saved_list.parts):
            saved_list.parts[index]['youtube_link'] = youtube_link
            saved_list.parts[index]['notes'] = notes
            saved_list.save()

            return jsonify({
                'success': True,
                'message': 'Part updated successfully'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Invalid part index'
            }), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/clear_all', methods=['POST'])
def clear_all():
    """Clear all saved parts"""
    try:
        saved_list.parts = []
        saved_list.save()
        return jsonify({
            'success': True,
            'message': 'All parts cleared!'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/export_excel')
def export_excel():
    """Export saved parts to Excel with clickable links"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import make_response

    # Create workbook with two sheets
    wb = Workbook()

    # Separate cars and trucks
    cars = [p for p in saved_list.parts if p.get('vehicle_type') == 'car']
    trucks = [p for p in saved_list.parts if p.get('vehicle_type') != 'car']

    def create_sheet(ws, parts, title):
        ws.title = title

        # Header style
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # Headers
        headers = ['Part Name', 'eBay Listing', 'Junkyard $', 'eBay $', 'ROI', 'Rating', 'YouTube', 'Notes', 'Added']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for row_idx, part in enumerate(parts, 2):
            ebay_price = part.get('ebay_price', part.get('ebay_sold_price', part.get('median_sold_price', 0)))
            roi_rating = part.get('roi_rating', 'N/A')

            # Part name
            ws.cell(row=row_idx, column=1, value=part.get('part_name', ''))

            # eBay listing (clickable link)
            ebay_url = part.get('ebay_url', '')
            ebay_title = part.get('ebay_title', 'View Listing')[:50]
            if ebay_url:
                cell = ws.cell(row=row_idx, column=2, value=ebay_title)
                cell.hyperlink = ebay_url
                cell.font = Font(color="0563C1", underline="single")
            else:
                ws.cell(row=row_idx, column=2, value=ebay_title)

            # Junkyard price
            ws.cell(row=row_idx, column=3, value=f"${part.get('junkyard_price', 0):.2f}")

            # eBay price
            ws.cell(row=row_idx, column=4, value=f"${ebay_price:.2f}")

            # ROI
            ws.cell(row=row_idx, column=5, value=f"{part.get('roi', 0):.2f}x")

            # ROI Rating with color
            rating_cell = ws.cell(row=row_idx, column=6, value=roi_rating)
            if roi_rating == 'High':
                rating_cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
                rating_cell.font = Font(color="FFFFFF", bold=True)
            elif roi_rating == 'Medium':
                rating_cell.fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
                rating_cell.font = Font(bold=True)
            else:
                rating_cell.fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
                rating_cell.font = Font(color="FFFFFF", bold=True)
            rating_cell.alignment = Alignment(horizontal='center')

            # YouTube link (clickable)
            youtube_url = part.get('youtube_link', '').strip()
            if youtube_url:
                cell = ws.cell(row=row_idx, column=7, value="Watch Video")
                cell.hyperlink = youtube_url
                cell.font = Font(color="FF0000", underline="single", bold=True)
            else:
                ws.cell(row=row_idx, column=7, value="-")

            # Notes
            ws.cell(row=row_idx, column=8, value=part.get('notes', '-'))

            # Date added
            ws.cell(row=row_idx, column=9, value=part.get('saved_at', ''))

        # Adjust column widths
        ws.column_dimensions['A'].width = 25  # Part Name
        ws.column_dimensions['B'].width = 40  # eBay Listing
        ws.column_dimensions['C'].width = 12  # Junkyard $
        ws.column_dimensions['D'].width = 12  # eBay $
        ws.column_dimensions['E'].width = 10  # ROI
        ws.column_dimensions['F'].width = 12  # Rating
        ws.column_dimensions['G'].width = 15  # YouTube
        ws.column_dimensions['H'].width = 35  # Notes
        ws.column_dimensions['I'].width = 18  # Added

    # Create Cars sheet
    if cars:
        ws_cars = wb.active
        create_sheet(ws_cars, cars, "Cars")

    # Create Trucks sheet
    if trucks:
        if not cars:
            ws_trucks = wb.active
        else:
            ws_trucks = wb.create_sheet("Trucks & SUVs")
        create_sheet(ws_trucks, trucks, "Trucks & SUVs")

    # Remove default sheet if we created custom ones
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb['Sheet'])

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=parts_list.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

@app.route('/export_csv')
def export_csv():
    """Export saved parts to CSV"""
    import csv
    from io import StringIO
    from flask import make_response

    output = StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(['Part Name', 'eBay Title', 'eBay URL', 'eBay Image', 'eBay Price', 'Junkyard Parts', 'Junkyard Price', 'ROI', 'ROI Rating', 'Vehicle Type', 'Year', 'Make', 'Model', 'YouTube Tutorial', 'Notes', 'Date Added'])

    # Write parts
    for part in saved_list.parts:
        writer.writerow([
            part.get('part_name', ''),
            part.get('ebay_title', ''),
            part.get('ebay_url', ''),
            part.get('best_listing_image', ''),  # eBay image URL
            part.get('ebay_price', part.get('ebay_sold_price', part.get('median_sold_price', 0))),
            ', '.join(part.get('junkyard_parts', [])) if part.get('junkyard_parts') else '',
            part.get('junkyard_price', 0),
            part.get('roi', 0),
            part.get('roi_rating', ''),
            part.get('vehicle_type', ''),
            part.get('year', ''),
            part.get('make', ''),
            part.get('model', ''),
            part.get('youtube_link', ''),
            part.get('notes', ''),
            part.get('saved_at', '')
        ])

    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=parts_list.csv'
    response.headers['Content-Type'] = 'text/csv'
    return response

@app.route('/download_html')
def download_html():
    """Download saved parts as HTML file"""
    from flask import make_response
    from datetime import datetime

    # Get the HTML content
    html_content = generate_parts_html()

    response = make_response(html_content)
    response.headers['Content-Disposition'] = 'attachment; filename=my-parts-list.html'
    response.headers['Content-Type'] = 'text/html'
    return response

@app.route('/export_html')
def export_html():
    """Export saved parts to mobile-friendly HTML"""
    from flask import make_response

    # Get the HTML content
    html_content = generate_parts_html()

    response = make_response(html_content)
    response.headers['Content-Type'] = 'text/html'
    return response

def generate_parts_html():
    """Generate mobile-friendly HTML for parts list"""
    from datetime import datetime

    # Separate cars and trucks
    cars = [p for p in saved_list.parts if p.get('vehicle_type') == 'car']
    trucks = [p for p in saved_list.parts if p.get('vehicle_type') != 'car']

    def create_parts_html(parts, title):
        if not parts:
            return ''

        html = f'<h2 style="color: #667eea; margin-top: 30px; border-bottom: 2px solid #667eea; padding-bottom: 10px;">{title}</h2>'

        for part in parts:
            ebay_price = part.get('ebay_price', part.get('ebay_sold_price', part.get('median_sold_price', 0)))
            junkyard_price = part.get('junkyard_price', 0)
            roi = part.get('roi', 0)
            roi_rating = part.get('roi_rating', 'N/A')

            # ROI badge color
            roi_color = '#28a745' if roi_rating == 'High' else ('#ffc107' if roi_rating == 'Medium' else '#dc3545')

            # Vehicle info
            vehicle_info = f"{part.get('year', '')} {part.get('make', '')} {part.get('model', '')}".strip()

            html += f'''
            <div style="background: white; padding: 20px; margin: 15px 0; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                <h3 style="margin: 0 0 15px 0; color: #333; font-size: 1.3em;">{part.get('part_name', 'N/A')}</h3>

                {f'<p style="margin: 5px 0; color: #666; font-size: 0.95em;"><strong>Vehicle:</strong> {vehicle_info}</p>' if vehicle_info else ''}

                <div style="margin: 15px 0; padding: 12px; background: #f0f9ff; border-left: 4px solid #667eea; border-radius: 4px;">
                    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
                        <p style="margin: 0; font-weight: bold; color: #667eea; font-size: 0.9em;">📋 EBAY TITLE:</p>
                        <button onclick="copyToClipboard('{part.get('ebay_title', 'N/A').replace("'", "\\'")}', this)" style="background: #667eea; color: white; border: none; padding: 6px 12px; border-radius: 6px; font-size: 0.85em; font-weight: bold; cursor: pointer;">📋 Copy</button>
                    </div>
                    <p style="margin: 0; padding: 8px; background: white; border-radius: 4px; font-family: monospace; font-size: 0.9em; color: #333; word-break: break-word;">
                        {part.get('ebay_title', 'N/A')}
                    </p>
                </div>

                <p style="margin: 10px 0;">
                    <a href="{part.get('ebay_url', '#')}" target="_blank" style="display: inline-block; background: #667eea; color: white; padding: 10px 20px; border-radius: 8px; text-decoration: none; font-weight: bold;">🔗 Open eBay Listing</a>
                </p>

                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin: 15px 0;">
                    <div style="background: #f8f9fa; padding: 12px; border-radius: 8px; text-align: center;">
                        <div style="font-size: 0.85em; color: #666;">Junkyard</div>
                        <div style="font-size: 1.4em; font-weight: bold; color: #28a745;">${junkyard_price:.2f}</div>
                    </div>
                    <div style="background: #f8f9fa; padding: 12px; border-radius: 8px; text-align: center;">
                        <div style="font-size: 0.85em; color: #666;">eBay Sold</div>
                        <div style="font-size: 1.4em; font-weight: bold; color: #667eea;">${ebay_price:.2f}</div>
                    </div>
                </div>

                <div style="text-align: center; margin: 15px 0;">
                    <div style="font-size: 0.85em; color: #666; margin-bottom: 5px;">ROI</div>
                    <div style="display: inline-block; background: {roi_color}; color: white; padding: 8px 20px; border-radius: 20px; font-size: 1.2em; font-weight: bold;">
                        {roi:.2f}x - {roi_rating}
                    </div>
                </div>

                {f'<p style="margin: 10px 0;"><a href="{part.get("youtube_link", "")}" target="_blank" style="display: inline-block; background: #ff0000; color: white; padding: 10px 20px; border-radius: 8px; text-decoration: none; font-weight: bold;">🎥 Watch Tutorial</a></p>' if part.get('youtube_link', '').strip() else ''}

                {f'<div style="background: #fff3cd; padding: 12px; border-radius: 8px; margin-top: 10px;"><strong style="color: #856404;">Notes:</strong><br><span style="color: #856404;">{part.get("notes", "")}</span></div>' if part.get('notes', '').strip() else ''}

                <p style="margin: 10px 0 0 0; font-size: 0.8em; color: #999;">Added: {part.get('saved_at', 'N/A')}</p>
            </div>
            '''

        return html

    html_content = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <title>Car Parts List - {datetime.now().strftime("%Y-%m-%d")}</title>
    <style>
        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 15px;
            line-height: 1.6;
        }}

        .container {{
            max-width: 800px;
            margin: 0 auto;
        }}

        h1 {{
            color: white;
            text-align: center;
            margin: 20px 0;
            font-size: 1.8em;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
        }}

        .summary {{
            background: white;
            padding: 20px;
            border-radius: 12px;
            margin-bottom: 20px;
            text-align: center;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }}

        .summary h3 {{
            color: #667eea;
            font-size: 1.5em;
            margin-bottom: 10px;
        }}

        @media print {{
            body {{
                background: white;
                padding: 0;
            }}

            .container {{
                max-width: 100%;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🚗 Car Parts List</h1>
        <div class="summary">
            <h3>Total Parts: {len(saved_list.parts)}</h3>
            <p style="color: #666; margin-top: 5px;">Cars: {len(cars)} | Trucks/SUVs: {len(trucks)}</p>
            <p style="color: #999; font-size: 0.9em; margin-top: 10px;">Exported: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}</p>
        </div>

        {create_parts_html(cars, '🚗 Cars')}
        {create_parts_html(trucks, '🚙 Trucks / SUVs')}
    </div>

    <script>
        function copyToClipboard(text, button) {{
            // Copy text to clipboard
            if (navigator.clipboard && navigator.clipboard.writeText) {{
                navigator.clipboard.writeText(text).then(function() {{
                    // Show success feedback
                    const originalText = button.innerHTML;
                    button.innerHTML = '✅ Copied!';
                    button.style.background = '#28a745';

                    // Reset button after 2 seconds
                    setTimeout(function() {{
                        button.innerHTML = originalText;
                        button.style.background = '#667eea';
                    }}, 2000);
                }}).catch(function(err) {{
                    alert('Failed to copy: ' + err);
                }});
            }} else {{
                // Fallback for older browsers/devices
                const textArea = document.createElement('textarea');
                textArea.value = text;
                textArea.style.position = 'fixed';
                textArea.style.opacity = '0';
                document.body.appendChild(textArea);
                textArea.select();
                try {{
                    document.execCommand('copy');
                    const originalText = button.innerHTML;
                    button.innerHTML = '✅ Copied!';
                    button.style.background = '#28a745';
                    setTimeout(function() {{
                        button.innerHTML = originalText;
                        button.style.background = '#667eea';
                    }}, 2000);
                }} catch (err) {{
                    alert('Failed to copy');
                }}
                document.body.removeChild(textArea);
            }}
        }}
    </script>
</body>
</html>'''

    return html_content

@app.route('/filter', methods=['POST'])
def filter_results():
    """Filter results by ROI or sort"""
    data = request.json
    results = data.get('results', [])
    filter_type = data.get('filter_type')
    min_roi = data.get('min_roi', 5.0)

    if filter_type == 'roi_filter':
        filtered = analyzer.filter_by_roi(results, min_roi)
    elif filter_type == 'sort_frequency':
        filtered = analyzer.sort_by_frequency(results)
    else:
        filtered = results

    return jsonify({
        'success': True,
        'results': filtered
    })

if __name__ == '__main__':
    # Validate configuration
    Config.validate()

    print("\n" + "="*60)
    print("eBay Car Parts Profit Analyzer")
    print("="*60)
    print(f"Junkyard parts loaded: {len(junkyard_prices.prices)}")
    print(f"Saved parts: {len(saved_list.parts)}")
    print("\nStarting web server at: http://localhost:5000")
    print("="*60 + "\n")

    app.run(debug=True, host='0.0.0.0', port=5000)
